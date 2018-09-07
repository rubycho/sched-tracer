# python 3.6

import os
import sys
import time
import config
import signal
import subprocess
import collections

from openpyxl import Workbook
import networkx as nx
import matplotlib.pyplot as plt

from config import *
from kernel4_13_16 import *

if len(sys.argv) != 3:
    print("{} <filename> <time>".format(sys.argv[0]))
    print("\t\tfilename: record name(create if doesn't exists)")
    print("\t\ttime: recording time")
    sys.exit(1)

if not os.path.exists(CAPTURE_DIR):
    os.makedirs(CAPTURE_DIR)

os.chdir(CAPTURE_DIR)

filename = sys.argv[1]
record_time = sys.argv[2]

# check root
# the subprocess will raise CalledProcessError if it fails
print_info("Is terminal on root?")
su = subprocess.run(sudo("ls"),
            shell=True,
            stdout=subprocess.PIPE,
            check=True
        )
su.stdout = ''
print_info("Root confirmed.")

isNew = True
ros_node_pids = []
ros_node_procs = []

if os.path.isfile("{}.dat".format(filename)) \
    and os.path.isfile("{}.dcd".format(filename)):
    print_warn("Opening recorded data...")
    isNew = False

# record trace
if isNew:
    p = subprocess.Popen(
        sudo("trace-cmd record -e sched* -o {}.dat".format(filename)),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        shell=True,
        preexec_fn=os.setpgrp
    )

    print_info("Recording for {} seconds...".format(record_time))
    time.sleep(int(record_time))
    os.system(sudo("kill -2 -{}".format(p.pid)))

    stdout = p.stdout.read().decode('utf-8')
    print_info("STDOUT ON RECORDING: ")
    print(stdout)

    stderr = p.stderr.read().decode('utf-8')
    print_info("STDERR on RECORDING: ")
    print(stderr)

    if not os.path.isfile("{}.dat".format(filename)):
        print_error("No record file. "
            + "Please check the stdout/stderr of trace-cmd printed above.")
        exit(1)

    isRos = input("Capture ROS Node PID? (Y/else)")
    if isRos == 'Y':
        print_info("Please do not terminate the target application...")

        p = subprocess.run(
            '../extra/pid.sh',
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            shell=True,
            check=True
        )

        stdout = p.stdout.decode('utf-8')
        for o in stdout.splitlines():
            pid = o.split()[1]
            p = subprocess.run(
                'ps -T -p ' + pid,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                shell=True,
                check=True
            )

            _stdout = p.stdout.decode('utf-8')
            for _o in _stdout.splitlines()[1:]:
                ros_node_pids.append(int(_o.split()[1]))
            ros_node_pids.append(-1)

        f = open('{}.txt'.format(filename), 'w')
        f.write(','.join(str(x) for x in ros_node_pids))
        f.close()

    print_info("Reading & Processing...")
    p = subprocess.run(
        "trace-cmd report {0}.dat > {0}.dcd".format(filename),
        stdout=subprocess.PIPE,
        check=True,
        shell=True
    )

with open("{}.dcd".format(filename)) as f:
    dcd = f.readlines()

if len(ros_node_pids) == 0 and os.path.isfile("{}.txt".format(filename)):
    with open("{}.txt".format(filename)) as f:
        tmp = f.read()
        ros_node_pids = [int(x) for x in tmp.split(',')]

switches = []
wakeups = []

# capture switch
process_list = []
for line in dcd:
    try:
        tmp = SchedSwitch(line)
        switches.append(tmp)
    except (EventUnmatchException, TooLessArgumentException) as e:
        pass

# capture wakeup
for line in dcd:
    try:
        tmp = SchedWakeup(line)
        wakeups.append(tmp)
    except (EventUnmatchException, TooLessArgumentException) as e:
        pass

for switch in switches:
    Process.apply_switch(process_list, switch)

for wakeup in wakeups:
    Process.apply_wakeup(process_list, wakeup)

print_info("The number of switches: {}".format(len(switches)))
print_info("The number of wakeups: {}".format(len(wakeups)))

wb = Workbook()
wb.remove(wb['Sheet'])

G = nx.DiGraph()

# sheet: list of process
process_sheet = wb.create_sheet('list')
process_sheet.freeze_panes = process_sheet['B2']
process_sheet['A1'] = 'process_name'
process_sheet['B1'] = 'pid'
process_sheet['C1'] = 'average_runtime(s)'
process_sheet['D1'] = 'dispersion_runtime'
process_sheet['E1'] = 'max_runtime(s)'
process_sheet['F1'] = 'min_runtime(s)'
process_sheet['G1'] = 'average_period(s)'
process_sheet['H1'] = 'dispersion_period'
process_sheet['I1'] = 'max_period(s)'
process_sheet['J1'] = 'min_period(s)'

process_list.sort(key=lambda x: x.get_pid())

shells = []
shell = []
for pid in ros_node_pids:
    if pid == -1:
        shells.append(shell)
        shell = []
    else:
        shell.append(pid)

for process in process_list:
    if process.get_pid() in ros_node_pids:
        ros_node_procs.append(process)
        G.add_node(str(process.get_pid()) + process.get_name())

    for idx, s in enumerate(shells):
        for _idx, _s in enumerate(s):
            if _s == process.get_pid():
                shells[idx][_idx] = process.get_name()

print(len(shells))

for idx, process in enumerate(process_list):
    if process.pid != 0:
        ws = wb.create_sheet(str(process.get_pid()))

        process_sheet['A' + str(idx+2)] = process.get_name()
        process_sheet['B' + str(idx+2)] = process.get_pid()
        process_sheet['C' + str(idx+2)] = process.get_runtime_average()
        process_sheet['D' + str(idx+2)] = process.get_runtime_dispersion()
        process_sheet['E' + str(idx+2)] = process.get_runtime_max()
        process_sheet['F' + str(idx+2)] = process.get_runtime_min()
        process_sheet['G' + str(idx+2)] = process.get_period_average()
        process_sheet['H' + str(idx+2)] = process.get_period_dispersion()
        process_sheet['I' + str(idx+2)] = process.get_period_max()
        process_sheet['J' + str(idx+2)] = process.get_period_min()

        for x in 'CDEFGHIJ':
            process_sheet[x + str(idx+2)].number_format = '#,##0.0000000000'

        ws.freeze_panes = ws['A2']
        ws['A1'] = process.get_name()
        ws['A2'] = 'in_stamp(s)'
        ws['B2'] = 'out_stamp(s)'
        ws['C2'] = 'runtime(s)'
        ws['D2'] = 'period(s)'
        ws['E2'] = 'waker'
        ws['F2'] = 'waker_count'
        ws['G2'] = 'wakee'
        ws['H2'] = 'wakee_count'

        try:
            first_in = process.get_in_stamp()[0]
            first_out = process.get_out_stamp()[0]
        except Exception:
            pass

        default = 3
        delay = 3
        if first_in and first_out and first_out < first_in:
            delay = 4

        for idx, item in enumerate(process.get_in_stamp()):
            ws['A' + str(idx+delay)] = item

        for idx, item in enumerate(process.get_out_stamp()):
            ws['B' + str(idx+default)] = item

        for idx, item in enumerate(process.get_runtime()):
            ws['C' + str(idx+delay)] = item

        for idx, item in enumerate(process.get_period()):
            ws['D' + str(idx+delay+1)] = item

        wakers = process.get_wakers()
        counter = collections.Counter(wakers)
        for idx, item in enumerate(counter.most_common()):
            ws['E' + str(idx+default)] = item[0]
            ws['F' + str(idx+default)] = item[1]

        wakees = process.get_wakees()
        counter = collections.Counter(wakees)
        for idx, item in enumerate(counter.most_common()):
            process_name = str(process.get_pid()) + process.get_name()

            tmp1 = item[0].split(':')[-1]
            tmp = tmp1 + item[0]

            if process_name in G.nodes() and tmp in G.nodes() and item[1] > 10:
                weight = item[1] / 25.0
                if weight < 1:
                    weight = 1
                if weight > 4:
                    weight = 4

                G.add_edge(process_name, tmp, weight=weight)

            ws['G' + str(idx+default)] = item[0]
            ws['H' + str(idx+default)] = item[1]

        for x in 'ABCD':
            for cell in ws[x]:
                cell.number_format = '#,##0.000000'

for sheet in wb.worksheets:
    for col in sheet.columns:
         max_length = 0
         column = col[0].column # Get the column name
         for cell in col:
             try: # Necessary to avoid error on empty cells
                 if len(str(cell.value)) > max_length:
                     max_length = len(cell.value)
             except:
                 pass
         adjusted_width = (max_length + 2) * 1.2
         sheet.column_dimensions[column].width = adjusted_width

wb.save(filename + '.xlsx')

pos = nx.nx_pydot.graphviz_layout(G, prog='neato')

labels = nx.get_edge_attributes(G, 'weight')
weights = [G[u][v]['weight'] for u,v in G.edges()]

nx.draw(G, pos, with_labels=True, width=weights)
plt.show()
